from openpyxl import *
from tkinter import *
from tkinter import ttk

wb = load_workbook('D:\\Bai_Tap\\Fuck\\Lab4 PY\\excel.xlsx')

sheet = wb.active


def excel():
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50
	sheet.cell(row=1, column=1).value = "Mã số sinh viên"
	sheet.cell(row=1, column=2).value = "Họ và tên"
	sheet.cell(row=1, column=3).value = "Ngày sinh"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "Số điện thoại"
	sheet.cell(row=1, column=6).value = "Học kì"
	sheet.cell(row=1, column=7).value = "Năm Học"

def focus1(event):
	course_field.focus_set()

def focus2(event):
	sem_field.focus_set()

def focus3(event):
	form_no_field.focus_set()

def focus4(event):
	contact_no_field.focus_set()

def focus5(event):
	email_id_field.focus_set()

def focus6(event):
	address_field.focus_set()
	
def clear():
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)
	
def insert():
	if (name_field.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):
			
		print("empty input")

	else:
		current_row = sheet.max_row
		current_column = sheet.max_column
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = course_field.get()
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
		sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
		sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
		sheet.cell(row=current_row + 1, column=7).value = address_field.get()

		# save the file
		wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx')

		name_field.focus_set()

		clear()


# Driver code
if __name__ == "__main__":
	
	# create a GUI window
	root = Tk()

	# set the background colour of GUI window
	root.configure(background='light green')

	# set the title of GUI window
	root.title("registration form")

	# set the configuration of GUI window
	root.geometry("500x300")

	excel()

	# create a Form label
	heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green")

	# create a Name label
	name = Label(root, text="Mã số sinh viên", bg="light green")

	# create a Course label
	course = Label(root, text="Họ và tên", bg="light green")

	# create a Semester label
	sem = Label(root, text="Ngày sinh", bg="light green")

	# create a Form No. label
	form_no = Label(root, text="Email", bg="light green")

	# create a Contact No. label
	contact_no = Label(root, text="Số điện thoại", bg="light green")

	# create a Email id label
	email_id = Label(root, text="Học kì", bg="light green")

	address = Label(root, text="Năm học", bg="light green")

	heading.grid(row=0, column=1)
	name.grid(row=1, column=0)
	course.grid(row=2, column=0)
	sem.grid(row=3, column=0)
	form_no.grid(row=4, column=0)
	contact_no.grid(row=5, column=0)
	email_id.grid(row=6, column=0)
	address.grid(row=7, column=0)

	name_field = Entry(root)
	course_field = Entry(root)
	sem_field = Entry(root)
	form_no_field = Entry(root)
	contact_no_field = Entry(root)
	email_id_field = Entry(root)
	address_field = Entry(root)


	name_field.bind("<Return>", focus1)

	course_field.bind("<Return>", focus2)

	sem_field.bind("<Return>", focus3)

	form_no_field.bind("<Return>", focus4)

	contact_no_field.bind("<Return>", focus5)
	email_id_field.bind("<Return>", focus6)
	name_field.grid(row=1, column=1, ipadx="100")
	course_field.grid(row=2, column=1, ipadx="100")
	sem_field.grid(row=3, column=1, ipadx="100")
	form_no_field.grid(row=4, column=1, ipadx="100")
	contact_no_field.grid(row=5, column=1, ipadx="100")
	email_id_field.grid(row=6, column=1, ipadx="100")
	address_field.grid(row=7, column=1, ipadx="100")

	excel()
	

	
	root.mainloop()
